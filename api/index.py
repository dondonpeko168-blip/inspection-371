"""Vercel serverless function for 371 inspection viewer - dynamic schema Flask version."""
import json
import os
import sqlite3
import gzip
import io
import boto3
from botocore.config import Config as BotoConfig
from flask import Flask, request, jsonify, Response, send_from_directory, make_response

app = Flask(__name__, static_folder="../")

DB_PATH = "/tmp/insp_371.db"
PAGE_SIZE = 100
ITEM_PAGE = 200

# ── R2 / S3 client ──────────────────────────────────────────────────────────
R2_ACCOUNT_ID  = os.environ.get("R2_ACCOUNT_ID",  "72ee811d77e8613491f208c9e92bb937")
R2_BUCKET      = os.environ.get("R2_BUCKET",      "371-file")
R2_ENDPOINT    = os.environ.get("R2_ENDPOINT",
                                  f"https://{R2_ACCOUNT_ID}.r2.cloudflarestorage.com")
R2_ACCESS_KEY  = os.environ.get("R2_ACCESS_KEY",  "")
R2_SECRET_KEY  = os.environ.get("R2_SECRET_KEY",  "")

_r2_client = None
def get_r2_client():
    global _r2_client
    if not _r2_client and R2_ACCESS_KEY and R2_SECRET_KEY:
        _r2_client = boto3.client(
            "s3",
            endpoint_url=R2_ENDPOINT,
            aws_access_key_id=R2_ACCESS_KEY,
            aws_secret_access_key=R2_SECRET_KEY,
            region_name="auto",
            config=BotoConfig(s3={"addressing_style": "path"}),
        )
    return _r2_client


def parse_xlsx(file_bytes):
    """Parse .xlsx bytes, return (headers: list[str], rows: list[list])"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else f"col_{i}"
               for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = [str(v) if v is not None else "" for v in row]
        # Pad or trim to match header count
        while len(r) < len(headers):
            r.append("")
        rows.append(r[:len(headers)])
    wb.close()
    return headers, rows


def ensure_db():
    """Build SQLite DB from data.xlsx.gz with dynamic schema."""
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_gz = os.path.join(script_dir, "data.xlsx.gz")
    xlsx_path = os.path.join(script_dir, "data.xlsx")

    if os.path.exists(xlsx_gz):
        with gzip.open(xlsx_gz, "rb") as f:
            xlsx_data = f.read()
    elif os.path.exists(xlsx_path):
        with open(xlsx_path, "rb") as f:
            xlsx_data = f.read()
    else:
        return None, "Data file not found", []

    headers, rows = parse_xlsx(xlsx_data)
    if not headers:
        return None, "No columns found in Excel", []

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA synchronous = OFF")
    conn.execute("PRAGMA journal_mode = MEMORY")
    conn.execute("PRAGMA cache_size = -16000")

    # Create table with dynamic column names (col_0, col_1, ...)
    col_names = [f"c{i}" for i in range(len(headers))]
    cols_def = ", ".join(f"{name} TEXT" for name in col_names)
    conn.execute(f"CREATE TABLE inspection ({cols_def})")

    placeholders = ",".join("?" * len(headers))
    batch_size = 5000
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        conn.executemany(f"INSERT INTO inspection VALUES ({placeholders})", batch)
        conn.commit()

    conn.execute("ANALYZE")
    # Store xlsx mtime for cache-busting on next deploy
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_gz_path = os.path.join(script_dir, "data.xlsx.gz")
    xlsx_mtime = int(os.path.getmtime(xlsx_gz_path) * 1000)
    conn.execute("CREATE TABLE IF NOT EXISTS _meta (key TEXT PRIMARY KEY, value TEXT)")
    conn.execute("INSERT OR REPLACE INTO _meta VALUES ('xlsx_mtime', ?)", (str(xlsx_mtime),))
    conn.commit()
    conn.close()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn, None, headers


def get_conn_or_error():
    """Reuse cached DB if it exists and is valid, otherwise rebuild."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    # Check if DB is valid by reading headers from file
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_gz = os.path.join(script_dir, "data.xlsx.gz")
    xlsx_path = os.path.join(script_dir, "data.xlsx")

    # Read the current xlsx headers
    if os.path.exists(xlsx_gz):
        with gzip.open(xlsx_gz, "rb") as f:
            xlsx_data = f.read()
    elif os.path.exists(xlsx_path):
        with open(xlsx_path, "rb") as f:
            xlsx_data = f.read()
    else:
        return None, jsonify({"error": "Data file not found"}), 500, []

    current_headers, _ = parse_xlsx(xlsx_data)

    # Check if DB already has the right schema
    try:
        cursor = conn.execute("PRAGMA table_info(inspection)")
        db_cols = [row[1] for row in cursor.fetchall()]
        # Force rebuild if column count doesn't match
        if len(db_cols) != len(current_headers):
            conn.close()
            if os.path.exists(DB_PATH):
                os.remove(DB_PATH)
            conn, err, headers = ensure_db()
            if err:
                return None, jsonify({"error": err}), 500, []
            return conn, None, None, headers
        # Check row count
        cursor = conn.execute("SELECT COUNT(*) FROM inspection")
        total = cursor.fetchone()[0]
        if total == 0:
            conn.close()
            if os.path.exists(DB_PATH):
                os.remove(DB_PATH)
            conn, err, headers = ensure_db()
            if err:
                return None, jsonify({"error": err}), 500, []
            return conn, None, None, headers
        # ── Force rebuild if xlsx file changed (new deployment) ──────────────
        xlsx_mtime = int(os.path.getmtime(xlsx_gz) * 1000)
        try:
            cursor = conn.execute("SELECT value FROM _meta WHERE key='xlsx_mtime'")
            row = cursor.fetchone()
            cached_mtime = int(row[0]) if row else 0
        except sqlite3.OperationalError:
            cached_mtime = 0
        if xlsx_mtime != cached_mtime:
            conn.close()
            if os.path.exists(DB_PATH):
                os.remove(DB_PATH)
            conn, err, headers = ensure_db()
            if err:
                return None, jsonify({"error": err}), 500, []
            return conn, None, None, headers
        return conn, None, None, current_headers
    except sqlite3.OperationalError:
        conn.close()
        if os.path.exists(DB_PATH):
            os.remove(DB_PATH)
        conn, err, headers = ensure_db()
        if err:
            return None, jsonify({"error": err}), 500, []
        return conn, None, None, headers


def row_to_dict(row, headers):
    """Convert a sqlite3.Row to a dict using dynamic headers."""
    d = {}
    for i, h in enumerate(headers):
        key = f"c{i}"
        d[h] = row[key] if key in row.keys() else ""
    return d


@app.before_request
def handle_OPTIONS():
    """Handle CORS preflight before Flask tries to route the OPTIONS request."""
    if request.method == "OPTIONS":
        resp = app.make_response("")
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
        return resp

@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response


@app.route("/api/debug", methods=["GET"])
def api_debug():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_gz = os.path.join(script_dir, "data.xlsx.gz")
    exists = os.path.exists(xlsx_gz)
    mtime = int(os.path.getmtime(xlsx_gz) * 1000) if exists else 0
    return jsonify({"xlsx_exists": exists, "xlsx_mtime": mtime})

@app.route("/api/init", methods=["GET", "OPTIONS"])
def api_init():
    conn, err_resp, err_code, headers = get_conn_or_error()
    if err_resp:
        return err_resp, err_code

    try:
        cursor = conn.execute("SELECT COUNT(*) as total FROM inspection")
        total = cursor.fetchone()["total"]
    except Exception:
        total = 0

    return jsonify({
        "total": total,
        "headers": headers,
        "page_size": PAGE_SIZE,
        "item_page": ITEM_PAGE
    })


@app.route("/api/items", methods=["GET", "OPTIONS"])
def api_items():
    conn, err_resp, err_code, headers = get_conn_or_error()
    if err_resp:
        return err_resp, err_code

    page = int(request.args.get("page", 1))
    offset = (page - 1) * ITEM_PAGE
    limit = ITEM_PAGE

    # Use first header column for item list (or c0 if headers exist)
    first_col = "c0"
    cursor = conn.execute(f"SELECT DISTINCT {first_col} FROM inspection LIMIT ? OFFSET ?", (limit, offset))
    items = [row[first_col] for row in cursor.fetchall()]

    pages = {"current": page, "items": items}
    return jsonify(pages)


@app.route("/api/query", methods=["GET", "OPTIONS"])
def api_query():
    conn, err_resp, err_code, headers = get_conn_or_error()
    if err_resp:
        return err_resp, err_code

    q = request.args.get("q", "").strip()
    page = int(request.args.get("page", 1))
    offset = (page - 1) * PAGE_SIZE
    limit = PAGE_SIZE

    col_names = [f"c{i}" for i in range(len(headers))]

    conditions = []
    params = []

    if q:
        like = f"%{q}%"
        ors = " OR ".join([f"{c} LIKE ?" for c in col_names])
        conditions.append(f"({ors})")
        params.extend([like] * len(col_names))

    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    cursor = conn.execute(f"SELECT COUNT(*) as total FROM inspection {where}", params)
    total = cursor.fetchone()["total"]

    query_sql = f"SELECT * FROM inspection {where} LIMIT ? OFFSET ?"
    params.append(limit)
    params.append(offset)
    cursor = conn.execute(query_sql, params)

    results = [row_to_dict(row, headers) for row in cursor.fetchall()]

    return jsonify({
        "total": total,
        "page": page,
        "per_page": limit,
        "data": results
    })


@app.route("/api/data", methods=["GET", "OPTIONS"])
def api_data():
    """Return all data records (browse endpoint)."""
    conn, err_resp, err_code, headers = get_conn_or_error()
    if err_resp:
        return err_resp, err_code

    limit = int(request.args.get("limit", 500))

    cursor = conn.execute(f"SELECT * FROM inspection LIMIT ?", (limit,))
    results = [row_to_dict(row, headers) for row in cursor.fetchall()]

    return jsonify(results)


@app.route("/api/export", methods=["POST", "OPTIONS"])
def api_export():
    conn, err_resp, err_code, headers = get_conn_or_error()
    if err_resp:
        return err_resp, err_code

    data = request.get_json() or {}
    text = data.get("text", "")
    export_format = data.get("format", "csv")

    col_names = [f"c{i}" for i in range(len(headers))]
    conditions = []
    params = []

    if text:
        like = f"%{text}%"
        ors = " OR ".join([f"{c} LIKE ?" for c in col_names])
        conditions.append(f"({ors})")
        params.extend([like] * len(col_names))

    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    cursor = conn.execute(f"SELECT * FROM inspection {where}", params)
    rows = cursor.fetchall()

    if export_format == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row[f"c{i}"] for i in range(len(headers))])
        result = output.getvalue()
        return Response(result, mimetype="text/csv")

    elif export_format == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        # Header row with bold
        header_font = Font(bold=True)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
        for row_idx, row in enumerate(rows, 2):
            for col_idx in range(len(headers)):
                ws.cell(row=row_idx, column=col_idx + 1,
                        value=row[f"c{col_idx}"])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(output.getvalue(),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return jsonify({"error": "Unsupported format"}), 400


# --- Upload: base64 via text/plain (CORS simple, no preflight) ---
@app.route("/api/upload-base64", methods=["POST"])
def api_upload_base64():
    """Accept base64 xlsx via text/plain body. text/plain is a CORS simple
    content-type → browser sends POST without OPTIONS preflight."""
    content_type = request.content_type or ""
    if "text/plain" in content_type:
        # Body is raw base64 string
        raw_body = request.get_data()
        if not raw_body:
            return jsonify({"error": "No data"}), 400
        try:
            import base64 as _b64
            file_bytes = _b64.b64decode(raw_body.decode("utf-8").strip())
        except Exception as e:
            return jsonify({"error": "Invalid base64: " + str(e)}), 400
        return _process_xlsx_bytes(file_bytes, "upload.xlsx")
    else:
        return jsonify({"error": "Expected Content-Type: text/plain"}), 400


def _process_xlsx_bytes(file_bytes, filename):
    """Parse xlsx bytes, write data.xlsx.gz, invalidate DB, return JSON."""
    import gzip, openpyxl, os as _os
    script_dir = _os.path.dirname(_os.path.abspath(__file__))
    xlsx_gz_path = _os.path.join(script_dir, "data.xlsx.gz")
    xlsx_path = _os.path.join(script_dir, "data.xlsx")
    db_path = "/tmp/insp_371.db"

    try:
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else f"col_{i}"
                   for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            r = [str(v) if v is not None else "" for v in row]
            while len(r) < len(headers): r.append("")
            rows.append(r[:len(headers)])
        wb.close()
        count = len(rows)
    except Exception as e:
        return jsonify({"error": "Failed to parse xlsx: " + str(e)}), 400

    # Write gzip
    try:
        with gzip.open(xlsx_gz_path, "wb", compresslevel=6) as gz:
            gz.write(file_bytes)
        with open(xlsx_path, "wb") as f:
            f.write(file_bytes)
        try: _os.remove(db_path)
        except: pass
    except Exception as e:
        return jsonify({"error": "File write error: " + str(e)}), 500

    sample = rows[:3]
    return jsonify({
        "success": True, "filename": filename,
        "rows": count, "columns": headers,
        "sample": [dict(zip(headers, r)) for r in sample],
    })


# --- Upload (new path, avoids any cached 501 from edge): handle multipart/form-data ---
@app.route("/api/upload-file", methods=["POST", "OPTIONS"])
def api_upload_file():
    if request.method == "OPTIONS":
        resp = make_response("")
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
        return resp
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_gz_path = os.path.join(script_dir, "data.xlsx.gz")
    xlsx_path = os.path.join(script_dir, "data.xlsx")

    # Handle multipart form-data
    if request.content_type and 'multipart/form-data' in request.content_type:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        file_bytes = file.read()
        filename = file.filename
    elif request.content_type and 'text/plain' in request.content_type:
        # Base64 text sent by browser (CORS simple, no preflight needed)
        raw_data = request.get_data()
        if not raw_data:
            return jsonify({"error": "No data"}), 400
        try:
            import base64 as _b64
            file_bytes = _b64.b64decode(raw_data.decode("utf-8").strip())
        except Exception as e:
            return jsonify({"error": "Invalid base64: " + str(e)}), 400
        filename = "upload.xlsx"
    else:
        raw_data = request.get_data()
        if not raw_data:
            return jsonify({"error": "No data"}), 400
        filename = request.args.get("filename", "upload.xlsx")
        if raw_data[:2] == b'\x1f\x8b':
            file_bytes = gzip.decompress(raw_data)
        else:
            file_bytes = raw_data

    if not filename.endswith('.xlsx') and not filename.endswith('.xlsx.gz'):
        return jsonify({"error": "Only .xlsx files are supported"}), 400

    # If uploaded as gzip, decompress
    if filename.endswith('.gz'):
        data_to_save = file_bytes  # keep as gzip
        xlsx_data = gzip.decompress(file_bytes)
    else:
        data_to_save = file_bytes
        xlsx_data = file_bytes

    # Parse and validate
    headers, rows = parse_xlsx(xlsx_data)
    if not headers:
        return jsonify({"error": "Failed to parse Excel"}), 400

    # Save as gzip
    if filename.endswith('.gz'):
        with open(xlsx_gz_path, "wb") as f:
            f.write(data_to_save)
    else:
        import gzip as gz_module
        with gz_module.open(xlsx_gz_path, "wb") as f:
            f.write(xlsx_data)

    # Also save .xlsx for fallback
    with open(xlsx_path, "wb") as f:
        f.write(xlsx_data)

    # Rebuild DB
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    conn, err, _ = ensure_db()
    if conn:
        conn.close()

    return jsonify({
        "success": True,
        "filename": filename,
        "rows": len(rows),
        "columns": headers,
        "sample": [{h: (rows[j][i] if i < len(rows[j]) else "") for i, h in enumerate(headers)}
                    for j in range(min(3, len(rows)))]
    })


# ── /api/upload: legacy endpoint — Vercel Edge needs explicit OPTIONS route ──
@app.route("/api/upload", methods=["GET", "POST", "OPTIONS"])
def api_upload():
    """Legacy upload endpoint. OPTIONS handled at edge by this Python route."""
    if request.method == "OPTIONS":
        resp = make_response("")
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        resp.headers["Access-Control-Max-Age"] = "86400"
        return resp
    # Redirect GET/POST to the proper handlers
    return jsonify({"message": "Use /api/upload-file for small files or /api/upload-url for large files"}), 200


# ── Presigned-URL Upload Flow (maintains backward compat) ────────────────────
@app.route("/api/upload-url", methods=["GET", "OPTIONS"])
def api_upload_url():
    """Return a presigned PUT URL for direct browser→R2 upload."""
    client = get_r2_client()
    if not client:
        return jsonify({"error": "Storage not configured"}), 503

    filename = request.args.get("filename", "upload.xlsx")
    content_type = request.args.get("content_type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Build a unique R2 key with timestamp to avoid collisions
    import time, uuid
    key = f"uploads/{int(time.time())}_{uuid.uuid4().hex[:8]}_{filename}"

    try:
        presigned = client.generate_presigned_url(
            "put_object",
            Params={
                "Bucket": R2_BUCKET,
                "Key": key,
                "ContentType": content_type,
            },
            ExpiresIn=3600,   # 1 hour
        )
        return jsonify({
            "upload_url": presigned,
            "key": key,
            "bucket": R2_BUCKET,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/upload-complete", methods=["POST", "OPTIONS"])
def api_upload_complete():
    """After browser PUTs to R2, call this to download & process the file."""
    client = get_r2_client()
    if not client:
        return jsonify({"error": "Storage not configured"}), 503

    data = request.get_json() or {}
    key = data.get("key", "").strip()
    filename = data.get("filename", "upload.xlsx")

    if not key:
        return jsonify({"error": "Missing 'key'"}), 400

    try:
        # Download from R2
        response = client.get_object(Bucket=R2_BUCKET, Key=key)
        file_bytes = response["Body"].read()

        # Decompress if needed
        if filename.endswith(".gz"):
            xlsx_data = gzip.decompress(file_bytes)
        else:
            xlsx_data = file_bytes

        # Parse & validate
        headers, rows = parse_xlsx(xlsx_data)
        if not headers:
            return jsonify({"error": "Failed to parse Excel"}), 400

        # Save as gzip locally
        script_dir = os.path.dirname(os.path.abspath(__file__))
        xlsx_gz_path = os.path.join(script_dir, "data.xlsx.gz")
        xlsx_path = os.path.join(script_dir, "data.xlsx")
        with gzip.open(xlsx_gz_path, "wb") as f:
            f.write(xlsx_data)
        with open(xlsx_path, "wb") as f:
            f.write(xlsx_data)

        # Rebuild DB
        if os.path.exists(DB_PATH):
            os.remove(DB_PATH)
        conn, err, _ = ensure_db()
        if conn:
            conn.close()

        # Clean up R2 object
        try:
            client.delete_object(Bucket=R2_BUCKET, Key=key)
        except Exception:
            pass  # non-fatal

        return jsonify({
            "success": True,
            "filename": filename,
            "rows": len(rows),
            "columns": headers,
            "sample": [{h: (rows[j][i] if i < len(rows[j]) else "") for i, h in enumerate(headers)}
                       for j in range(min(3, len(rows)))],
        })
    except client.exceptions.NoSuchKey:
        return jsonify({"error": "File not found in storage"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def catch_all(path):
    """Serve index.html for all non-API routes."""
    return send_from_directory("../", "index.html")


# ── Filter + Export: Upload Excel → Filter by keywords → Download ─────────────
@app.route("/api/filter-upload", methods=["POST", "OPTIONS"])
@app.route("/api/filter", methods=["POST", "OPTIONS"])
def api_filter_upload():
    """Accept Excel file + keywords, filter rows, return CSV/XLSX download.
    Works for files up to ~50MB. Uses streaming where possible."""
    if request.method == "OPTIONS":
        resp = make_response("")
        resp.headers["Access-Control-Allow-Origin"] = "https://inspection-371.vercel.app"
        resp.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        resp.headers["Access-Control-Max-Age"] = "86400"
        return resp

    # ── Parse form data ─────────────────────────────────────────────────────
    if request.content_type and 'multipart/form-data' in request.content_type:
        file = request.files.get('file')
        keywords = request.form.get('keywords', '').strip()
        export_format = request.form.get('format', 'csv').lower()
    else:
        data = request.get_json(silent=True) or {}
        file = None
        keywords = data.get('keywords', '')
        export_format = data.get('format', 'csv').lower()

    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    if not keywords:
        return jsonify({"error": "請輸入關鍵字（keywords）"}), 400
    if export_format not in ('csv', 'xlsx'):
        return jsonify({"error": "format 僅支援 csv 或 xlsx"}), 400

    try:
        file_bytes = file.read()
    except Exception as e:
        return jsonify({"error": f"讀取檔案失敗: {e}"}), 400

    # ── Parse Excel ─────────────────────────────────────────────────────────
    try:
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        hdr_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(hdr_row)]
        keyword_list = [k.strip() for k in keywords.split(",") if k.strip()]
        filtered = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_text = " ".join(str(v) for v in row if v is not None).lower()
            if all(kw.lower() in row_text for kw in keyword_list):
                filtered.append([str(v) if v is not None else "" for v in row])
        wb.close()
    except Exception as e:
        return jsonify({"error": f"Excel 解析失敗: {e}"}), 400

    if not filtered:
        return jsonify({"error": "沒有找到符合的資料"}), 404

    # ── Export ──────────────────────────────────────────────────────────────
    total_rows = len(filtered)

    if export_format == "csv":
        import csv as _csv
        output = _io.StringIO()
        writer = _csv.writer(output)
        writer.writerow(headers)
        writer.writerows(filtered)
        result_bytes = output.getvalue().encode("utf-8-sig")
        return Response(
            result_bytes,
            mimetype="text/csv; charset=utf-8-sig",
            headers={
                "Content-Disposition": f"attachment; filename=filtered_{len(filtered)}_rows.csv",
                "Content-Length": str(len(result_bytes)),
                "X-Filtered-Count": str(total_rows),
            }
        )
    else:  # xlsx
        import openpyxl
        from openpyxl.styles import Font
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        hdr_font = Font(bold=True)
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = hdr_font
        for ri, row in enumerate(filtered, 2):
            for ci, val in enumerate(row, 1):
                ws2.cell(row=ri, column=ci, value=val)
        buf = _io.BytesIO()
        wb2.save(buf)
        buf.seek(0)
        result_bytes = buf.getvalue()
        return Response(
            result_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=filtered_{len(filtered)}_rows.xlsx",
                "Content-Length": str(len(result_bytes)),
                "X-Filtered-Count": str(total_rows),
            }
        )


# Vercel entry point
handler = app.wsgi_app  # force rebuild
