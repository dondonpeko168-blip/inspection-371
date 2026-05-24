from http.server import BaseHTTPRequestHandler
import json
import os
import sys

def parse_excel_bytes(file_bytes):
    """Parse .xlsx file and return list of records."""
    try:
        import openpyxl
        from io import BytesIO
    except ImportError:
        return None, "openpyxl not available"

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {}
            for i, h in enumerate(headers):
                if h is not None:
                    key = str(h).strip()
                    if key:
                        record[key] = row[i] if i < len(row) else None
            if any(v is not None for v in record.values()):
                records.append(record)
        return records, None
    except Exception as e:
        return None, str(e)

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps({
            "message": "Upload endpoint. POST a multipart/form-data with 'file' field."
        }).encode())

    def do_POST(self):
        import cgi
        from io import BytesIO
        
        content_type = self.headers.get('Content-Type', '')
        if 'multipart/form-data' not in content_type:
            self.send_response(400)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({"error": "Expected multipart/form-data"}).encode())
            return

        try:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={'REQUEST_METHOD': 'POST', 'CONTENT_TYPE': content_type}
            )
            
            if 'file' not in form:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"error": "No file uploaded"}).encode())
                return
            
            file_item = form['file']
            file_bytes = file_item.file.read()
            filename = file_item.filename or 'unknown.xlsx'
            
            if not filename.endswith('.xlsx'):
                self.send_response(400)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"error": "Only .xlsx files are supported"}).encode())
                return
            
            records, error = parse_excel_bytes(file_bytes)
            if error:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"error": f"Failed to parse Excel: {error}"}).encode())
                return
            
            columns = []
            if records:
                columns = list(records[0].keys())
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({
                "success": True,
                "filename": filename,
                "rows": len(records),
                "columns": columns,
                "sample": records[:3] if records else []
            }, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({"error": f"Server error: {str(e)}"}).encode())
